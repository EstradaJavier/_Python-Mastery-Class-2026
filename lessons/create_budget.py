import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, Reference
from openpyxl.drawing.image import Image
import datetime

# Create workbook
wb = Workbook()

# ==================== SHEET 1: Summary_Dashboard ====================
ws1 = wb.active
ws1.title = "Summary_Dashboard"

ws1['A1'] = "Estrada Family Budget - Monthly Summary"
ws1['A1'].font = Font(bold=True, size=14)

# Table Headers
ws1['A2'] = "Metric"
ws1['B2'] = "Planned_Amount"
for cell in ['A2', 'B2']:
    ws1[cell].font = Font(bold=True)
    ws1[cell].fill = PatternFill(start_color="B0E0E6", end_color="B0E0E6", fill_type="solid")

# Data
data = [
    ["Total Monthly Income", 11635],
    ["Total Fixed Bills", 4803],
    ["Total Variable Spending", 1650],
    ["Debt Payments", 450],
    ["Planned Savings/Buffer", 800],
    ["Projected Surplus/Deficit", 3932]
]

for r_idx, row in enumerate(data, start=3):
    ws1.cell(row=r_idx, column=1, value=row[0])
    ws1.cell(row=r_idx, column=2, value=row[1])

# Color Legend
ws1['A11'] = "Category_Type"
ws1['B11'] = "Color_Explanation"
ws1['A11'].font = Font(bold=True)
ws1['B11'].font = Font(bold=True)

legend = [
    ["Essential", "Must be paid (mortgage, utilities, basic food, meds)"],
    ["Flexible", "Important but adjustable (groceries, basic pets, gas)"],
    ["Discretionary", "Nice-to-have (Amazon, eating out, premium subscriptions)"],
    ["Savings/Buffer", "Money reserved for emergencies or future"]
]

for r_idx, row in enumerate(legend, start=12):
    ws1.cell(row=r_idx, column=1, value=row[0])
    ws1.cell(row=r_idx, column=2, value=row[1])

# Pie Chart
chart = PieChart()
labels = Reference(ws1, min_col=1, min_row=4, max_row=7)
data_ref = Reference(ws1, min_col=2, min_row=4, max_row=7)
chart.add_data(data_ref, titles_from_data=False)
chart.set_categories(labels)
chart.title = "Monthly Plan - Spending Breakdown"
ws1.add_chart(chart, "E2")

# ==================== SHEET 2: Monthly_Budget_Plan ====================
ws2 = wb.create_sheet("Monthly_Budget_Plan")
ws2['A1'] = "Monthly Budget Plan - Duplicate this sheet for each month"
ws2['A1'].font = Font(bold=True, size=12)

headers = ["Category", "Type (Income/Fixed/Variable/Debt/Savings)", "Due_Date", "Planned_Amount", "Account", "Notes"]
for col, header in enumerate(headers, start=1):
    cell = ws2.cell(row=2, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="B0E0E6", end_color="B0E0E6", fill_type="solid")

# ==================== SHEET 3: Cash_Flow_Calendar ====================
ws3 = wb.create_sheet("Cash_Flow_Calendar")
ws3['A1'] = "Cash Flow Calendar - One Month View"
ws3['A1'].font = Font(bold=True, size=12)

headers3 = ["Date", "Starting_Balance", "Income", "Bills/Spending", "Ending_Balance", "Notes"]
for col, header in enumerate(headers3, start=1):
    cell = ws3.cell(row=2, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="B0E0E6", end_color="B0E0E6", fill_type="solid")

# Sample dates and formula
for i in range(3, 35):
    ws3.cell(row=i, column=1, value=f"2026-05-{i-2:02d}")
    if i == 3:
        ws3.cell(row=i, column=2, value=5400)  # Starting balance example
    ws3.cell(row=i, column=5, value=f"=B{i} + C{i} - D{i}")

# ==================== SHEET 4: Actual_Transactions ====================
ws4 = wb.create_sheet("Actual_Transactions")
ws4['A1'] = "Actual Transactions - Categorize each line"
ws4['A1'].font = Font(bold=True, size=12)

headers4 = ["Date", "Description", "Amount", "Category", "Account", "Notes"]
for col, header in enumerate(headers4, start=1):
    cell = ws4.cell(row=2, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="B0E0E6", end_color="B0E0E6", fill_type="solid")

# ==================== SHEET 5: Category_Legend ====================
ws5 = wb.create_sheet("Category_Legend")
ws5['A1'] = "Categories & Color Legend"
ws5['A1'].font = Font(bold=True, size=12)

headers5 = ["Category", "Group", "Color_Suggestion", "Notes"]
for col, header in enumerate(headers5, start=1):
    cell = ws5.cell(row=2, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="B0E0E6", end_color="B0E0E6", fill_type="solid")

legend_data = [
    ["Mortgage", "Essential", "Green", "House payment"],
    ["Utilities", "Essential", "Green", "Water, T-Mobile, Starlink"],
    ["Groceries", "Flexible", "Yellow", "Planned grocery budget"],
    ["Dining Out", "Discretionary", "Red", "Restaurants, DoorDash"],
    ["Amazon/Online", "Discretionary", "Red", "Amazon & online shopping"],
    ["Gas/Transport", "Essential", "Green", "Fuel"],
    ["Pet Care", "Flexible", "Yellow", "Dogs & 4 Cats"],
    ["Medical/Pharmacy", "Essential", "Green", "Copays, prescriptions"],
    ["Subscriptions", "Discretionary", "Red", "Streaming services"],
    ["Debt Payments", "Essential", "Green", "Credit cards"],
    ["Miscellaneous", "Flexible", "Yellow", "Everything else"],
    ["Emergency Savings", "Savings", "Blue", "Do not touch"]
]

for r_idx, row in enumerate(legend_data, start=3):
    for c_idx, value in enumerate(row, start=1):
        ws5.cell(row=r_idx, column=c_idx, value=value)

# Save the file
file_name = "Estrada_Family_Budget_Template.xlsx"
wb.save(file_name)
print(f"Workbook successfully created: {file_name}")